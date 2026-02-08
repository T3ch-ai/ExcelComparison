"""
Data loader: Loads QES from Excel or CSV files, NIQ from Excel / RDS / CSV / mock.
All column names come from config.
Supports chunked loading for large provider files (1GB+).
"""

import os
import pandas as pd
from config_loader import get_rds_connection_string


def load_qes_data(cfg: dict) -> tuple:
    """Load QES-set-1 (Network Adequacy) and QES-set-2 (Providers).
    Supports Excel (.xlsx/.xls) and CSV (.csv) files.
    Format is determined by config source_type or auto-detected from file extension."""
    qes = cfg["qes"]
    state = cfg["state"]
    state_col = cfg["key_columns"]["qes"][0]
    chunk_cfg = cfg.get("chunked_loading", {})

    wb1_path = qes["workbook1_path"]
    wb2_path = qes["workbook2_path"]

    # QES Network Adequacy (set-1) -- typically small
    qes_na = _read_file(wb1_path, qes.get("workbook1_sheet", 0), qes.get("source_type"))
    qes_na = qes_na[qes_na[state_col] == state].reset_index(drop=True)

    # QES Providers (set-2) -- potentially large
    source_type = _detect_source_type(wb2_path, qes.get("source_type"))
    if source_type == "csv" and _should_chunk(wb2_path, chunk_cfg):
        qes_provs = _load_csv_chunked(wb2_path, state_col, state, chunk_cfg)
    elif source_type == "excel" and _should_chunk(wb2_path, chunk_cfg):
        qes_provs = _load_excel_chunked(
            wb2_path, qes.get("workbook2_sheet", 0), state_col, state, chunk_cfg
        )
    else:
        qes_provs = _read_file(wb2_path, qes.get("workbook2_sheet", 0), qes.get("source_type"))
        qes_provs = qes_provs[qes_provs[state_col] == state].reset_index(drop=True)

    print(f"  [loaded] QES-set-1: {len(qes_na)} rows from {wb1_path}")
    print(f"  [loaded] QES-set-2: {len(qes_provs)} rows from {wb2_path}")
    return qes_na, qes_provs


def _detect_source_type(file_path: str, configured_type: str = None) -> str:
    """Determine file format from config or file extension."""
    if configured_type and configured_type in ("excel", "csv"):
        return configured_type
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return "csv"
    return "excel"


def _read_file(file_path: str, sheet_name=0, source_type: str = None) -> pd.DataFrame:
    """Read a single file as Excel or CSV based on source_type or file extension."""
    fmt = _detect_source_type(file_path, source_type)
    if fmt == "csv":
        return pd.read_csv(file_path)
    return pd.read_excel(file_path, sheet_name=sheet_name)


def load_niq_data(cfg: dict) -> tuple:
    """Load NIQ data based on configured source_type."""
    src = cfg["niq"]["source_type"]
    loaders = {
        "excel": _load_niq_excel,
        "rds": _load_niq_rds,
        "csv": _load_niq_csv,
        "mock": _load_niq_mock,
    }
    if src not in loaders:
        raise ValueError(f"Unknown niq.source_type: {src}")
    return loaders[src](cfg)


def _load_niq_excel(cfg: dict) -> tuple:
    """Load NIQ from Excel workbook(s). Supports single workbook with two sheets
    or two separate workbooks -- controlled by config."""
    excel_cfg = cfg["niq"]["excel"]
    state = cfg["state"]
    state_col = cfg["key_columns"]["niq"][0]
    chunk_cfg = cfg.get("chunked_loading", {})

    single = excel_cfg.get("single_workbook", True)

    if single:
        wb_path = excel_cfg["workbook_path"]
        na_sheet = excel_cfg.get("network_adequacy_sheet", 0)
        prov_sheet = excel_cfg.get("provider_detail_sheet", 1)
        niq_na = pd.read_excel(wb_path, sheet_name=na_sheet)

        # Provider sheet may be large
        if _should_chunk(wb_path, chunk_cfg):
            niq_provs = _load_excel_chunked(wb_path, prov_sheet, state_col, state, chunk_cfg)
        else:
            niq_provs = pd.read_excel(wb_path, sheet_name=prov_sheet)

        src_label = f"{wb_path} [{na_sheet}, {prov_sheet}]"
    else:
        wb1 = excel_cfg["workbook1_path"]
        wb2 = excel_cfg["workbook2_path"]
        niq_na = pd.read_excel(wb1, sheet_name=excel_cfg.get("workbook1_sheet", 0))

        if _should_chunk(wb2, chunk_cfg):
            niq_provs = _load_excel_chunked(
                wb2, excel_cfg.get("workbook2_sheet", 0), state_col, state, chunk_cfg
            )
        else:
            niq_provs = pd.read_excel(wb2, sheet_name=excel_cfg.get("workbook2_sheet", 0))

        src_label = f"{wb1} + {wb2}"

    niq_na = niq_na[niq_na[state_col] == state].reset_index(drop=True)
    if state_col in niq_provs.columns:
        niq_provs = niq_provs[niq_provs[state_col] == state].reset_index(drop=True)

    print(f"  [loaded] NIQ-set-1: {len(niq_na)} rows from {src_label}")
    print(f"  [loaded] NIQ-set-2: {len(niq_provs)} rows from {src_label}")
    return niq_na, niq_provs


def _load_niq_rds(cfg: dict) -> tuple:
    from sqlalchemy import create_engine, text

    conn_str = get_rds_connection_string(cfg)
    engine = create_engine(conn_str)
    rds_cfg = cfg["niq"]["rds"]
    state = cfg["state"]

    with engine.connect() as conn:
        niq_na = pd.read_sql(text(rds_cfg["network_adequacy_query"]), conn, params={"state": state})
        niq_provs = pd.read_sql(text(rds_cfg["provider_detail_query"]), conn, params={"state": state})

    print(f"  [loaded] NIQ-set-1: {len(niq_na)} rows from RDS")
    print(f"  [loaded] NIQ-set-2: {len(niq_provs)} rows from RDS")
    return niq_na, niq_provs


def _load_niq_csv(cfg: dict) -> tuple:
    csv_cfg = cfg["niq"]["csv"]
    state = cfg["state"]
    state_col = cfg["key_columns"]["niq"][0]
    chunk_cfg = cfg.get("chunked_loading", {})

    # Network adequacy file (typically small)
    niq_na = pd.read_csv(csv_cfg["network_adequacy_path"])
    niq_na = niq_na[niq_na[state_col] == state].reset_index(drop=True)

    # Provider file (potentially large)
    prov_path = csv_cfg["provider_detail_path"]
    if _should_chunk(prov_path, chunk_cfg):
        niq_provs = _load_csv_chunked(prov_path, state_col, state, chunk_cfg)
    else:
        niq_provs = pd.read_csv(prov_path)
        niq_provs = niq_provs[niq_provs[state_col] == state].reset_index(drop=True)

    print(f"  [loaded] NIQ-set-1: {len(niq_na)} rows from CSV")
    print(f"  [loaded] NIQ-set-2: {len(niq_provs)} rows from CSV")
    return niq_na, niq_provs


def _load_niq_mock(cfg: dict) -> tuple:
    from mock_data_generator import generate_all_data, save_qes_workbooks, save_niq_workbook

    data = generate_all_data(cfg)
    save_qes_workbooks(data, cfg)
    save_niq_workbook(data, cfg)

    print(f"  [loaded] NIQ-set-1: {len(data['niq_na'])} rows (mock)")
    print(f"  [loaded] NIQ-set-2: {len(data['niq_providers'])} rows (mock)")
    return data["niq_na"], data["niq_providers"]


# ============================================================================
# Chunked Loading Utilities
# ============================================================================

def _should_chunk(file_path: str, chunk_cfg: dict) -> bool:
    """Determine if file should be loaded in chunks based on size."""
    if not chunk_cfg.get("enabled", False):
        return False
    threshold_mb = chunk_cfg.get("provider_file_threshold_mb", 100)
    try:
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        return file_size_mb > threshold_mb
    except OSError:
        return False


def _load_csv_chunked(file_path: str, filter_col: str, filter_val: str, chunk_cfg: dict) -> pd.DataFrame:
    """Load a large CSV file in chunks, filtering as we go to minimize memory."""
    chunk_size = chunk_cfg.get("chunk_size", 50000)
    chunks = []
    total_read = 0
    for chunk in pd.read_csv(file_path, chunksize=chunk_size):
        total_read += len(chunk)
        if filter_col in chunk.columns:
            filtered = chunk[chunk[filter_col] == filter_val]
            if len(filtered) > 0:
                chunks.append(filtered)
        if total_read % (chunk_size * 10) == 0:
            print(f"    ... processed {total_read:,} rows")

    if chunks:
        result = pd.concat(chunks, ignore_index=True)
    else:
        result = pd.DataFrame()

    print(f"  [loaded] Provider data: {len(result)} rows from {total_read:,} total (chunked CSV)")
    return result


def _load_excel_chunked(file_path: str, sheet_name, filter_col: str, filter_val: str, chunk_cfg: dict) -> pd.DataFrame:
    """Stream-read an Excel file using openpyxl read_only mode for large files."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows_iter))]

    if filter_col not in headers:
        wb.close()
        # Fallback: load normally if filter column not found in headers
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        return df

    filter_idx = headers.index(filter_col)
    chunk_size = chunk_cfg.get("chunk_size", 50000)

    chunks = []
    buffer = []
    total_read = 0

    for row_values in rows_iter:
        total_read += 1
        if str(row_values[filter_idx]) == str(filter_val):
            buffer.append(row_values)

        if len(buffer) >= chunk_size:
            chunks.append(pd.DataFrame(buffer, columns=headers))
            buffer = []

        if total_read % (chunk_size * 10) == 0:
            print(f"    ... processed {total_read:,} rows")

    if buffer:
        chunks.append(pd.DataFrame(buffer, columns=headers))

    wb.close()

    result = pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame(columns=headers)
    print(f"  [loaded] Provider data: {len(result)} rows from {total_read:,} total (chunked Excel)")
    return result
