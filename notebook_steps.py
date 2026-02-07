# %% [markdown]
# # Network Adequacy Comparison -- Step-by-Step Notebook
#
# Run each cell independently. Inspect intermediate DataFrames and test
# individual components before running the full pipeline.
#
# **VSCode**: This `.py` file uses `# %%` markers. VSCode treats each block
# as a Jupyter cell. Click "Run Cell" above any block, or open as a notebook
# via right-click -> "Open as Jupyter Notebook".
#
# **Setup**: Run `./setup.sh` first to create the venv and install packages.

# %% [markdown]
# ---
# ## 1. Setup and Imports

# %%
import sys, os

PROJECT_DIR = os.path.dirname(os.path.abspath("__file__"))
if PROJECT_DIR not in sys.path:
    sys.path.insert(0, PROJECT_DIR)
os.chdir(PROJECT_DIR)

# Load .env if python-dotenv is available
try:
    from dotenv import load_dotenv
    load_dotenv()
    print("[OK] .env loaded")
except ImportError:
    print("[SKIP] python-dotenv not installed, .env not loaded")

print(f"Working directory: {os.getcwd()}")

# %%
# Verify dependencies
import importlib

for pkg in ["pandas", "openpyxl", "yaml"]:
    try:
        importlib.import_module(pkg)
        print(f"  [OK] {pkg}")
    except ImportError:
        print(f"  [MISSING] {pkg}")

# %%
# Import project modules
from config_loader import load_config
from data_loader import load_qes_data, load_niq_data
from comparator import compare_network_adequacy
from workbook_builder import build_output_workbook
from mock_data_generator import generate_all_data, save_qes_workbooks, save_niq_workbook

import pandas as pd
import yaml

print("[OK] All project modules imported")

# %% [markdown]
# ---
# ## 2. Load and Inspect Configuration

# %%
cfg = load_config("config.yaml")

print(f"State            : {cfg['state']}")
print(f"NIQ Source       : {cfg['niq']['source_type']}")
print(f"QES Key Columns  : {cfg['key_columns']['qes']}")
print(f"NIQ Key Columns  : {cfg['key_columns']['niq']}")

# %%
# Compare columns detail
print("Compare Columns:")
print("-" * 65)
for cc in cfg["compare_columns"]:
    tol = cc.get("tolerance", 0)
    print(f"  {cc['label']:30s}  QES: {cc['qes_col']:25s}  NIQ: {cc['niq_col']:25s}  dtype={cc.get('dtype','text')}  tol={tol}")

# %%
# Additional (non-compared) columns
print("Additional Result Columns:")
print("-" * 65)
for ac in cfg.get("additional_result_columns", []):
    qes = ac.get("qes_col") or "(none)"
    niq = ac.get("niq_col") or "(none)"
    print(f"  {ac['label']:30s}  QES: {qes:25s}  NIQ: {niq}")

# %%
# Result column order
print("Result Column Order:")
for i, token in enumerate(cfg.get("result_column_order", []), 1):
    print(f"  {i}. {token}")

# %%
# Output labels (should all be plain ASCII)
print("Output Labels:")
for k, v in cfg["output"]["labels"].items():
    has_unicode = any(ord(c) > 127 for c in v)
    flag = " <-- UNICODE!" if has_unicode else ""
    print(f"  {k:25s} = {v}{flag}")

# %% [markdown]
# ---
# ## 3. Generate Mock Data (skip if using real files)
#
# This creates:
# - `input/QES_Network_Adequacy.xlsx` (QES-set-1)
# - `input/QES_Providers.xlsx` (QES-set-2)
# - `input/NIQ_Data.xlsx` with sheets `NetworkAdequacy` + `ProviderDetail`

# %%
os.makedirs("input", exist_ok=True)
os.makedirs("output", exist_ok=True)

data = generate_all_data(cfg)

print("Generated DataFrames:")
for name, df in data.items():
    print(f"  {name:20s} : {df.shape[0]:5d} rows x {df.shape[1]} cols")

# %%
# Inspect QES Network Adequacy (set-1) -- first 10 rows
print("QES-set-1 columns:", list(data["qes_na"].columns))
data["qes_na"].head(10)

# %%
# Inspect QES Providers (set-2) -- first 10 rows
print("QES-set-2 columns:", list(data["qes_providers"].columns))
data["qes_providers"].head(10)

# %%
# Inspect NIQ Network Adequacy (set-1) -- note different column names
print("NIQ-set-1 columns:", list(data["niq_na"].columns))
data["niq_na"].head(10)

# %%
# Inspect NIQ Providers (set-2)
print("NIQ-set-2 columns:", list(data["niq_providers"].columns))
data["niq_providers"].head(10)

# %%
# Save mock data to Excel files
save_qes_workbooks(data, cfg)
save_niq_workbook(data, cfg)
print("\nInput files written. Check input/ directory.")

# %% [markdown]
# ---
# ## 4. Load Data via data_loader (same path main.py takes)
#
# This tests the loader module end-to-end, reading from files on disk.

# %%
# Load QES from Excel workbooks
qes_na, qes_providers = load_qes_data(cfg)

print(f"\nQES-set-1 shape: {qes_na.shape}")
print(f"QES-set-2 shape: {qes_providers.shape}")

# %%
qes_na.head()

# %%
qes_providers.head()

# %%
# Load NIQ (source_type from config: mock/excel/csv/rds)
niq_na, niq_providers = load_niq_data(cfg)

print(f"\nNIQ-set-1 shape: {niq_na.shape}")
print(f"NIQ-set-2 shape: {niq_providers.shape}")

# %%
niq_na.head()

# %%
niq_providers.head()

# %% [markdown]
# ### 4a. Quick sanity checks on loaded data

# %%
# Verify key columns exist in both datasets
print("Key column check:")
for qk, nk in zip(cfg["key_columns"]["qes"], cfg["key_columns"]["niq"]):
    qes_ok = qk in qes_na.columns
    niq_ok = nk in niq_na.columns
    status = "OK" if (qes_ok and niq_ok) else "FAIL"
    print(f"  [{status}] QES '{qk}' ({qes_ok})  |  NIQ '{nk}' ({niq_ok})")

# %%
# Verify compare columns exist
print("Compare column check:")
for cc in cfg["compare_columns"]:
    qes_ok = cc["qes_col"] in qes_na.columns
    niq_ok = cc["niq_col"] in niq_na.columns
    status = "OK" if (qes_ok and niq_ok) else "FAIL"
    print(f"  [{status}] {cc['label']:30s}  QES '{cc['qes_col']}' ({qes_ok})  |  NIQ '{cc['niq_col']}' ({niq_ok})")

# %%
# Verify additional columns exist
print("Additional column check:")
for ac in cfg.get("additional_result_columns", []):
    qes_ok = ac.get("qes_col") is None or ac["qes_col"] in qes_na.columns
    niq_ok = ac.get("niq_col") is None or ac["niq_col"] in niq_na.columns
    status = "OK" if (qes_ok and niq_ok) else "FAIL"
    qes_col = ac.get("qes_col") or "(none)"
    niq_col = ac.get("niq_col") or "(none)"
    print(f"  [{status}] {ac['label']:30s}  QES '{qes_col}' ({qes_ok})  |  NIQ '{niq_col}' ({niq_ok})")

# %%
# Unique key counts -- helps spot duplicates or missing combos
qes_key_cols = cfg["key_columns"]["qes"]
niq_key_cols = cfg["key_columns"]["niq"]

qes_keys = set(qes_na[qes_key_cols].astype(str).agg("|".join, axis=1))
niq_keys = set(niq_na[niq_key_cols].astype(str).agg("|".join, axis=1))

print(f"Unique keys in QES : {len(qes_keys)}")
print(f"Unique keys in NIQ : {len(niq_keys)}")
print(f"In both            : {len(qes_keys & niq_keys)}")
print(f"QES only           : {len(qes_keys - niq_keys)}")
print(f"NIQ only           : {len(niq_keys - qes_keys)}")

# %% [markdown]
# ---
# ## 5. Run Comparison

# %%
comparison = compare_network_adequacy(qes_na, niq_na, cfg)

print(f"\nComparison result shape: {comparison.shape}")
print(f"Columns: {list(comparison.columns)}")

# %%
# Full comparison table
comparison

# %%
# Filter: only mismatched rows
labels = cfg["output"]["labels"]
mismatched = comparison[comparison["Overall_Match"] == labels["overall_mismatch"]]
print(f"Mismatched rows: {len(mismatched)}")
mismatched

# %%
# Filter: QES-only rows
qes_only = comparison[comparison["Row_Source"] == "QES Only"]
print(f"QES Only rows: {len(qes_only)}")
qes_only

# %%
# Filter: NIQ-only rows
niq_only = comparison[comparison["Row_Source"] == "NIQ Only"]
print(f"NIQ Only rows: {len(niq_only)}")
niq_only

# %%
# Mismatch breakdown by compare column
print("Mismatches per column:")
print("-" * 40)
for cc in cfg["compare_columns"]:
    match_col = f"Match_{cc['label']}"
    if match_col in comparison.columns:
        mm = len(comparison[comparison[match_col] == labels["mismatch"]])
        total = len(comparison[comparison["Row_Source"] == "Both"])
        print(f"  {cc['label']:30s} : {mm}/{total}")

# %%
# Inspect a specific mismatch in detail (first mismatched row)
if len(mismatched) > 0:
    row = mismatched.iloc[0]
    print("First mismatched row detail:")
    print("=" * 50)
    for col in comparison.columns:
        print(f"  {col:40s} : {row[col]}")
else:
    print("No mismatches found.")

# %% [markdown]
# ---
# ## 6. Build Output Workbook

# %%
output_path = build_output_workbook(
    qes_na, qes_providers,
    niq_na, niq_providers,
    comparison, cfg,
)

print(f"\nOutput written to: {output_path}")

# %%
# Verify output workbook structure
from openpyxl import load_workbook

wb = load_workbook(output_path)
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"\nMain sheets:")
main_sheets = ["Summary", "Comparison_Results", "QES_Network_Adequacy",
                "QES_Providers", "NIQ_Network_Adequacy", "NIQ_Providers"]
for name in main_sheets:
    if name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name:30s}  {ws.max_row:5d} rows x {ws.max_column} cols")

dd_sheets = [s for s in wb.sheetnames if s not in main_sheets]
print(f"\nDrill-down sheets: {len(dd_sheets)}")
if dd_sheets:
    print(f"  First 5: {dd_sheets[:5]}")

# %%
# Verify hyperlinks in Comparison_Results
ws = wb["Comparison_Results"]
link_count = 0
print("Sample hyperlinks:")
for row in ws.iter_rows(min_row=2, max_row=min(8, ws.max_row)):
    for cell in row:
        if cell.hyperlink:
            link_count += 1
            if link_count <= 6:
                col_name = ws.cell(row=1, column=cell.column).value
                print(f"  Row {cell.row}, Col '{col_name}': value={cell.value} -> {cell.hyperlink.target}")
print(f"\nTotal hyperlinks found (first 7 data rows): {link_count}")

# %%
# Verify no unicode anywhere in Comparison_Results
ws = wb["Comparison_Results"]
unicode_cells = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        val = str(cell.value) if cell.value else ""
        if any(ord(c) > 127 for c in val):
            unicode_cells.append((cell.coordinate, val))

if unicode_cells:
    print(f"[FAIL] Unicode found in {len(unicode_cells)} cells:")
    for coord, val in unicode_cells[:10]:
        print(f"  {coord}: {repr(val)}")
else:
    print("[PASS] No unicode characters in Comparison_Results")

wb.close()

# %% [markdown]
# ---
# ## 7. Component Tests
#
# Isolated tests for individual functions. Useful when making changes.

# %% [markdown]
# ### 7a. Test config_loader

# %%
# Test: valid config loads without error
test_cfg = load_config("config.yaml")
assert "state" in test_cfg
assert "compare_columns" in test_cfg
assert len(test_cfg["key_columns"]["qes"]) == len(test_cfg["key_columns"]["niq"])
print("[PASS] config_loader.load_config")

# %%
# Test: labels are all ASCII
for k, v in test_cfg["output"]["labels"].items():
    assert all(ord(c) <= 127 for c in v), f"Unicode in label '{k}': {v}"
print("[PASS] All labels are ASCII")

# %% [markdown]
# ### 7b. Test mock_data_generator

# %%
mock_data = generate_all_data(cfg)

# Verify all 4 DataFrames are non-empty
for name in ["qes_na", "qes_providers", "niq_na", "niq_providers"]:
    assert len(mock_data[name]) > 0, f"{name} is empty"
    print(f"[PASS] {name}: {len(mock_data[name])} rows")

# Verify column names match config expectations
for qk in cfg["key_columns"]["qes"]:
    assert qk in mock_data["qes_na"].columns, f"Missing QES key col: {qk}"
for nk in cfg["key_columns"]["niq"]:
    assert nk in mock_data["niq_na"].columns, f"Missing NIQ key col: {nk}"
print("[PASS] Key columns present in mock data")

# Verify additional columns exist
for ac in cfg.get("additional_result_columns", []):
    if ac.get("qes_col"):
        assert ac["qes_col"] in mock_data["qes_na"].columns, f"Missing QES col: {ac['qes_col']}"
    if ac.get("niq_col"):
        assert ac["niq_col"] in mock_data["niq_na"].columns, f"Missing NIQ col: {ac['niq_col']}"
print("[PASS] Additional columns present in mock data")

# %% [markdown]
# ### 7c. Test comparator edge cases

# %%
# Test: identical datasets should produce 100% match
from comparator import _compare_values

lbls = cfg["output"]["labels"]

# Numeric exact match
match, diff = _compare_values(10, 10, "numeric", 0, lbls)
assert match is True and diff == 0
print("[PASS] Numeric exact match")

# Numeric within tolerance
match, diff = _compare_values(10.0, 10.005, "numeric", 0.01, lbls)
assert match is True
print(f"[PASS] Numeric within tolerance (diff={diff})")

# Numeric outside tolerance
match, diff = _compare_values(10.0, 10.05, "numeric", 0.01, lbls)
assert match is False
print(f"[PASS] Numeric outside tolerance (diff={diff})")

# Text match (case insensitive)
match, diff = _compare_values("Y", "y", "text", 0, lbls)
assert match is True
print("[PASS] Text case-insensitive match")

# Text mismatch
match, diff = _compare_values("Y", "N", "text", 0, lbls)
assert match is False
print(f"[PASS] Text mismatch (diff='{diff}')")

# Null handling
match, diff = _compare_values(None, None, "numeric", 0, lbls)
assert match is True
print("[PASS] Both null = match")

match, diff = _compare_values(10, None, "numeric", 0, lbls)
assert match is False
print("[PASS] One null = mismatch")

# %%
# Test: full comparison with small synthetic DataFrames
small_qes = pd.DataFrame({
    "State": ["TX", "TX", "TX"],
    "County_Name": ["Harris", "Dallas", "Travis"],
    "Specialty": ["Cardiology", "Cardiology", "Cardiology"],
    "Provider_Count": [10, 20, 30],
    "Meets_Standard": ["Y", "Y", "N"],
    "Avg_Distance_Miles": [5.0, 10.0, 35.0],
    "Member_Count": [1000, 2000, 3000],
})

small_niq = pd.DataFrame({
    "state_code": ["TX", "TX", "TX"],
    "county_name": ["Harris", "Dallas", "Bexar"],        # Bexar not in QES, Travis not in NIQ
    "specialty_type": ["Cardiology", "Cardiology", "Cardiology"],
    "provider_cnt": [10, 22, 15],                         # Dallas count differs
    "meets_standard_flag": ["Y", "Y", "Y"],
    "avg_distance": [5.0, 10.0, 8.0],
    "member_count": [1000, 2000, 4000],
})

small_result = compare_network_adequacy(small_qes, small_niq, cfg)
print(f"\nSmall test result: {len(small_result)} rows")
small_result

# %%
# Verify small test results
assert len(small_result) == 4, f"Expected 4 rows, got {len(small_result)}"

harris = small_result[small_result["County_Name"] == "Harris"].iloc[0]
assert harris["Overall_Match"] == labels["overall_match"], "Harris should match"

dallas = small_result[small_result["County_Name"] == "Dallas"].iloc[0]
assert dallas["Overall_Match"] == labels["overall_mismatch"], "Dallas should mismatch (count 20 vs 22)"
assert dallas["Diff_Provider Count"] == 2, f"Dallas diff should be 2, got {dallas['Diff_Provider Count']}"

travis = small_result[small_result["County_Name"] == "Travis"].iloc[0]
assert travis["Row_Source"] == "QES Only"

bexar = small_result[small_result["County_Name"] == "Bexar"].iloc[0]
assert bexar["Row_Source"] == "NIQ Only"

print("[PASS] All small-test assertions passed")

# %% [markdown]
# ### 7d. Test column ordering

# %%
from comparator import _apply_column_order

# Verify result_column_order from config is applied
ordered_cols = list(small_result.columns)
print("Column order in result:")
for i, col in enumerate(ordered_cols, 1):
    print(f"  {i:2d}. {col}")

# Verify keys come first
key_cols = cfg["key_columns"]["qes"]
for i, kc in enumerate(key_cols):
    assert ordered_cols[i] == kc, f"Position {i} should be '{kc}', got '{ordered_cols[i]}'"
print("\n[PASS] Key columns are first")

# Verify Overall_Match is last
assert ordered_cols[-1] == "Overall_Match"
print("[PASS] Overall_Match is last")

# %% [markdown]
# ### 7e. Test data_loader with Excel source

# %%
# Switch to excel source to test that path
import copy

excel_cfg = copy.deepcopy(cfg)
excel_cfg["niq"]["source_type"] = "excel"

niq_na_excel, niq_prov_excel = load_niq_data(excel_cfg)

print(f"NIQ-set-1 from Excel: {niq_na_excel.shape}")
print(f"NIQ-set-2 from Excel: {niq_prov_excel.shape}")

# Should match what mock generated
assert len(niq_na_excel) == len(niq_na), \
    f"Row count mismatch: excel={len(niq_na_excel)} vs mock={len(niq_na)}"
print("[PASS] Excel loader returns same row count as mock")

# %% [markdown]
# ---
# ## 8. Exploratory Analysis
#
# Ad-hoc queries once you have real data loaded.

# %%
# Distribution of provider counts across counties
print("QES Provider Count by County:")
print(qes_na.groupby(cfg["key_columns"]["qes"][1])["Provider_Count"].agg(["mean", "min", "max", "sum"]).round(1))

# %%
# Counties with the most mismatches
if len(mismatched) > 0:
    county_col = cfg["key_columns"]["qes"][1]
    print("Mismatches by County:")
    print(mismatched[county_col].value_counts())
else:
    print("No mismatches to analyze.")

# %%
# Provider overlap check for a specific county+specialty
# (useful when counts match but providers differ)
sample_county = qes_na[cfg["key_columns"]["qes"][1]].iloc[0]
sample_spec = qes_na[cfg["key_columns"]["qes"][2]].iloc[0]

qes_npis = set(
    qes_providers[
        (qes_providers["County_Name"] == sample_county) &
        (qes_providers["Specialty"] == sample_spec)
    ]["Provider_NPI"].astype(str)
)
niq_npis = set(
    niq_providers[
        (niq_providers["county_name"] == sample_county) &
        (niq_providers["specialty_type"] == sample_spec)
    ]["provider_npi"].astype(str)
)

print(f"Provider NPI overlap for {sample_county} / {sample_spec}:")
print(f"  QES providers : {len(qes_npis)}")
print(f"  NIQ providers : {len(niq_npis)}")
print(f"  In both       : {len(qes_npis & niq_npis)}")
print(f"  QES only      : {len(qes_npis - niq_npis)}")
print(f"  NIQ only      : {len(niq_npis - qes_npis)}")

# %% [markdown]
# ---
# ## 9. Override Config and Re-run
#
# Example: change state, add tolerance, test with different settings.

# %%
# Override example: increase tolerance for Avg Distance
import copy

custom_cfg = copy.deepcopy(cfg)

# Widen distance tolerance to 5.0
for cc in custom_cfg["compare_columns"]:
    if cc["label"] == "Avg Distance (Miles)":
        cc["tolerance"] = 5.0
        print(f"Set tolerance for '{cc['label']}' to {cc['tolerance']}")

# Re-run comparison with new tolerance
comparison_v2 = compare_network_adequacy(qes_na, niq_na, custom_cfg)

# Compare mismatch counts
orig_mm = len(comparison[comparison["Overall_Match"] == labels["overall_mismatch"]])
new_mm = len(comparison_v2[comparison_v2["Overall_Match"] == labels["overall_mismatch"]])
print(f"\nMismatches with default tolerance : {orig_mm}")
print(f"Mismatches with tolerance=5.0     : {new_mm}")
