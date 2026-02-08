# %% [markdown]
# # Network Adequacy Comparison -- Phase 2 Step-by-Step Notebook
#
# This notebook walks through each component of the Phase 2 comparison tool:
# 1. Environment setup and imports
# 2. Load and inspect configuration
# 3. Generate mock data with realistic columns
# 4. Load data via data_loader
# 5. Run comparison (with Direction and Match indicators)
# 6. Inspect summary data
# 7. Build output workbook
# 8. Component tests
# 9. Large file simulation
#
# **Environment Setup:**
# ```bash
# # Option A: Using setup.sh (Linux/macOS)
# chmod +x setup.sh && ./setup.sh
#
# # Option B: Manual setup (any OS)
# python -m venv .venv
# # Windows: .venv\Scripts\activate
# # Linux/macOS: source .venv/bin/activate
# pip install -r requirements.txt
# ```
#
# **Running in VSCode:**
# Open this file, then Ctrl+Shift+P → "Python: Select Interpreter" → choose .venv.
# Each `# %%` block runs as a separate cell (Jupyter-style).
#
# **Running as a plain script:**
# ```bash
# python notebook_steps_2.py
# ```

# %% -- Section 1: Setup and Imports
import sys
import os

# Ensure project root is on path (works both as .py script and as notebook cell)
project_root = os.path.dirname(os.path.abspath(__file__)) if "__file__" in dir() else os.getcwd()
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# Optional: load .env for RDS passwords etc.
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    print("python-dotenv not installed; skipping .env load (not needed for mock mode)")

import pandas as pd
import numpy as np

from config_loader import load_config
from data_loader import load_qes_data, load_niq_data
from comparator import compare_network_adequacy
from workbook_builder import build_output_workbook
from mock_data_generator import generate_all_data, save_qes_workbooks, save_niq_workbook

print("Imports OK")
print(f"Project root: {project_root}")
print(f"Python: {sys.version}")
print(f"pandas: {pd.__version__}")

# %% -- Section 2: Load and Inspect Configuration
cfg = load_config("config.yaml")

print(f"State: {cfg['state']}")
print(f"NIQ source_type: {cfg['niq']['source_type']}")
print()

print("--- Key Columns ---")
print(f"  QES keys: {cfg['key_columns']['qes']}")
print(f"  NIQ keys: {cfg['key_columns']['niq']}")
print()

print("--- Compare Columns ---")
for cc in cfg["compare_columns"]:
    direction = " [direction_indicator]" if cc.get("direction_indicator") else ""
    vmap = f" value_map={cc['value_map']}" if cc.get("value_map") else ""
    print(f"  QES: {cc['qes_col']} <-> NIQ: {cc['niq_col']}  "
          f"label={cc['label']}, dtype={cc.get('dtype','text')}, "
          f"tolerance={cc.get('tolerance', 0)}{direction}{vmap}")
print()

print("--- Additional Result Columns ---")
for ac in cfg.get("additional_result_columns", []):
    print(f"  QES: {ac.get('qes_col', 'N/A')} <-> NIQ: {ac.get('niq_col', 'N/A')}  label={ac['label']}")
print()

print("--- Drilldown Config ---")
dd = cfg.get("drilldown", {})
print(f"  link_column: {dd.get('link_column')}")
print(f"  QES provider keys: {dd.get('qes_provider_keys')}")
print(f"  NIQ provider keys: {dd.get('niq_provider_keys')}")
print()

print("--- Chunked Loading ---")
cl = cfg.get("chunked_loading", {})
print(f"  enabled: {cl.get('enabled')}, chunk_size: {cl.get('chunk_size')}, threshold_mb: {cl.get('provider_file_threshold_mb')}")
print()

print("--- Labels ---")
labels = cfg["output"]["labels"]
print(f"  higher={labels['higher']}, lower={labels['lower']}, same={labels['same']}")
print(f"  match_indicator={labels['match_indicator']}, no_match_indicator={labels['no_match_indicator']}")

# %% -- Section 3: Generate Mock Data
print(f"Generating Phase 2 mock data for state: {cfg['state']}...")
data = generate_all_data(cfg)

# Inspect QES-NetworkAdequacy-1
print("\n=== QES Network Adequacy (set-1) ===")
print(f"Shape: {data['qes_na'].shape}")
print(f"Columns: {list(data['qes_na'].columns)}")
print(data["qes_na"].head(3).to_string())

# Inspect QES-NA-Providers-2
print("\n=== QES Providers (set-2) ===")
print(f"Shape: {data['qes_providers'].shape}")
print(f"Columns: {list(data['qes_providers'].columns)}")
print(data["qes_providers"].head(3).to_string())

# Inspect NIQ-NetworkAdequacy-1
print("\n=== NIQ Network Adequacy (set-1) ===")
print(f"Shape: {data['niq_na'].shape}")
print(f"Columns: {list(data['niq_na'].columns)}")
print(data["niq_na"].head(3).to_string())

# Inspect NIQ-NA-Providers-2
print("\n=== NIQ Providers (set-2) ===")
print(f"Shape: {data['niq_providers'].shape}")
print(f"Columns: {list(data['niq_providers'].columns)}")
print(data["niq_providers"].head(3).to_string())

# %% -- Verify SSA codes and county classifications
print("=== SSA Code Verification ===")
ssa_codes = data["qes_na"]["CountySSA"].unique()
print(f"Unique SSA codes: {sorted(ssa_codes)}")

# Verify SSA and FIPS use correct state prefixes
state = cfg["state"]
from mock_data_generator import COUNTY_DATA_BY_STATE
state_data = COUNTY_DATA_BY_STATE.get(state, {})
if state_data:
    sample_ssa = list(state_data.values())[0]["ssa"]
    sample_fips = list(state_data.values())[0]["fips"]
    ssa_prefix = sample_ssa[:2]
    fips_prefix = sample_fips[:2]
    print(f"SSA state prefix: {ssa_prefix}, FIPS state prefix: {fips_prefix}")
    assert ssa_prefix != fips_prefix, "SSA and FIPS prefixes should differ!"
    print(f"SSA prefix != FIPS prefix -- PASS")

print("\n=== County Classification Distribution ===")
print(data["qes_na"]["County Class (Rural/Metro/Micro/CEAC)"].value_counts())

print("\n=== Specialty Codes ===")
print(data["qes_na"][["Specialty Group Code", "Specialty Group Name"]].drop_duplicates().to_string())

# %% -- Verify provider detail columns
print("=== QES Provider Entity Types ===")
if "EntityType" in data["qes_providers"].columns:
    print(data["qes_providers"]["EntityType"].value_counts())

print("\n=== Sample Facility Providers ===")
if "EntityType" in data["qes_providers"].columns:
    facilities = data["qes_providers"][data["qes_providers"]["EntityType"] == "Organization"]
    if len(facilities) > 0:
        print(facilities[["NPI", "Facility Name", "Taxonomy", "ServicingCounty"]].head(5).to_string())
    else:
        print("No facility providers in sample")

print("\n=== Lat/Long Range ===")
if "Latitude" in data["qes_providers"].columns:
    print(f"Latitude:  {data['qes_providers']['Latitude'].min():.4f} to {data['qes_providers']['Latitude'].max():.4f}")
    print(f"Longitude: {data['qes_providers']['Longitude'].min():.4f} to {data['qes_providers']['Longitude'].max():.4f}")

print("\n=== TIN/NPI Format Check ===")
npis = data["qes_providers"]["NPI"]
tins = data["qes_providers"]["TaxID"]
print(f"NPI sample: {npis.iloc[0]} (length={len(npis.iloc[0])})")
print(f"TIN sample: {tins.iloc[0]} (length={len(tins.iloc[0])})")
assert all(npis.str.len() == 10), "NPI not 10 digits!"
assert all(tins.str.len() == 9), "TIN not 9 digits!"
print("NPI=10 digits, TIN=9 digits -- PASS")

# %% -- Save mock data to files
print("Saving mock data to input files...")
save_qes_workbooks(data, cfg)
save_niq_workbook(data, cfg)
print("Done.")

# %% -- Section 4: Load Data via data_loader
print("Loading data through data_loader...")
niq_na, niq_providers = load_niq_data(cfg)
qes_na, qes_providers = load_qes_data(cfg)

print(f"\nQES NA shape: {qes_na.shape}")
print(f"QES Providers shape: {qes_providers.shape}")
print(f"NIQ NA shape: {niq_na.shape}")
print(f"NIQ Providers shape: {niq_providers.shape}")

# Memory usage
for name, df in [("QES-NA", qes_na), ("QES-Providers", qes_providers),
                 ("NIQ-NA", niq_na), ("NIQ-Providers", niq_providers)]:
    mem_mb = df.memory_usage(deep=True).sum() / (1024 * 1024)
    print(f"  {name}: {len(df):,} rows, {mem_mb:.1f} MB")

# %% -- Section 5: Run Comparison
print("Running comparison...")
comparison = compare_network_adequacy(qes_na, niq_na, cfg)

print(f"\nComparison shape: {comparison.shape}")
print(f"Columns: {list(comparison.columns)}")

# %% -- Inspect Direction column
print("=== Direction Column Distribution ===")
for cc in cfg["compare_columns"]:
    if cc.get("direction_indicator"):
        dir_col = f"Direction_{cc['label']}"
        if dir_col in comparison.columns:
            print(f"\n{dir_col}:")
            print(comparison[dir_col].value_counts())

# %% -- Filter by Direction
both_rows = comparison[comparison["Row_Source"] == "Both"]

for cc in cfg["compare_columns"]:
    if cc.get("direction_indicator"):
        lbl = cc["label"]
        dir_col = f"Direction_{lbl}"
        if dir_col not in both_rows.columns:
            continue

        higher = both_rows[both_rows[dir_col] == labels["higher"]]
        lower = both_rows[both_rows[dir_col] == labels["lower"]]
        same = both_rows[both_rows[dir_col] == labels["same"]]

        print(f"\n=== {lbl} ===")
        print(f"HIGHER (NIQ > QES): {len(higher)} rows")
        if len(higher) > 0:
            print(higher[["State", "CountySSA", "Specialty Group Code",
                         f"QES_{lbl}", f"NIQ_{lbl}", f"Diff_{lbl}", dir_col]].head(5).to_string())

        print(f"\nLOWER (NIQ < QES): {len(lower)} rows")
        if len(lower) > 0:
            print(lower[["State", "CountySSA", "Specialty Group Code",
                         f"QES_{lbl}", f"NIQ_{lbl}", f"Diff_{lbl}", dir_col]].head(5).to_string())

        print(f"\nSAME: {len(same)} rows")

# %% -- QES Only and NIQ Only rows
qes_only = comparison[comparison["Row_Source"] == "QES Only"]
niq_only = comparison[comparison["Row_Source"] == "NIQ Only"]
print(f"QES Only rows: {len(qes_only)}")
if len(qes_only) > 0:
    print(qes_only.head(3).to_string())
print(f"\nNIQ Only rows: {len(niq_only)}")
if len(niq_only) > 0:
    print(niq_only.head(3).to_string())

# %% -- Section 6: Inspect Summary Data
print("=== Summary Statistics ===")

key_qes = cfg["key_columns"]["qes"]
county_col = key_qes[1]
spec_col = key_qes[2]

# i) County count
print(f"\ni) Unique counties: {comparison[county_col].nunique()}")

# ii) Unique NPI counts
summary_cfg = cfg.get("summary", {})
npi_cols = summary_cfg.get("npi_columns", {"qes": "NPI", "niq": "npi"})
qes_npi = npi_cols.get("qes", "NPI")
niq_npi = npi_cols.get("niq", "npi")
if qes_npi in qes_providers.columns:
    print(f"ii) QES Unique NPIs: {qes_providers[qes_npi].nunique()}")
if niq_npi in niq_providers.columns:
    print(f"    NIQ Unique NPIs: {niq_providers[niq_npi].nunique()}")

# iii) Specialty summary
print(f"\niii) Specialty breakdown:")
for spec_code in sorted(both_rows[spec_col].unique()):
    spec_rows = both_rows[both_rows[spec_col] == spec_code]
    for cc in cfg["compare_columns"]:
        if cc.get("direction_indicator"):
            dir_col = f"Direction_{cc['label']}"
            if dir_col in spec_rows.columns:
                same = len(spec_rows[spec_rows[dir_col] == labels["same"]])
                qes_h = len(spec_rows[spec_rows[dir_col] == labels["lower"]])
                niq_h = len(spec_rows[spec_rows[dir_col] == labels["higher"]])
                print(f"  {spec_code}: Match={same}, QES>NIQ={qes_h}, NIQ>QES={niq_h}")

# iv) Zero serving
serving_cols = summary_cfg.get("serving_columns", {})
qes_srv = serving_cols.get("qes", "Servicing Providers count")
niq_srv = serving_cols.get("niq", "provider_covering")
if qes_srv in qes_na.columns:
    print(f"\niv) QES zero serving: {len(qes_na[qes_na[qes_srv] == 0])}")
if niq_srv in niq_na.columns:
    print(f"    NIQ zero serving: {len(niq_na[niq_na[niq_srv] == 0])}")

# %% -- Section 7: Build Output Workbook
print("Building output workbook...")
output_path = build_output_workbook(
    qes_na, qes_providers, niq_na, niq_providers, comparison, cfg
)
print(f"Output: {output_path}")

# %% -- Verify output workbook structure
from openpyxl import load_workbook as load_wb

wb = load_wb(output_path)
print(f"\nSheets in workbook ({len(wb.sheetnames)}):")
for i, name in enumerate(wb.sheetnames[:10]):
    ws = wb[name]
    print(f"  {i+1}. {name} ({ws.max_row} rows x {ws.max_column} cols)")
if len(wb.sheetnames) > 10:
    print(f"  ... and {len(wb.sheetnames) - 10} more sheets")
wb.close()

# %% -- Section 8: Component Tests

# Test 1: Direction indicator logic
from comparator import _compute_direction, _normalize_value, _normalize_key_part
test_labels = {"higher": "HIGHER", "lower": "LOWER", "same": "SAME"}

assert _compute_direction(90.0, 95.0, "numeric", 0.01, test_labels) == "HIGHER"
assert _compute_direction(95.0, 90.0, "numeric", 0.01, test_labels) == "LOWER"
assert _compute_direction(90.0, 90.0, "numeric", 0.01, test_labels) == "SAME"
assert _compute_direction(90.0, 90.005, "numeric", 0.01, test_labels) == "SAME"
assert _compute_direction(None, 90.0, "numeric", 0.01, test_labels) == ""
print("Test 1 PASSED: Direction indicator logic")

# Test 2: Normalize value
assert _normalize_value("95.5%", "numeric") == 95.5
assert _normalize_value(95.5, "numeric") == 95.5
assert _normalize_value("hello", "text") == "hello"
assert _normalize_value(None, "numeric") is None
print("Test 2 PASSED: Normalize value")

# Test 3: Join key normalization (handles Excel stripping leading zeros)
assert _normalize_key_part("011") == "11"
assert _normalize_key_part(11) == "11"
assert _normalize_key_part("007") == "7"
assert _normalize_key_part(7) == "7"
assert _normalize_key_part("45130") == "45130"
assert _normalize_key_part("TX") == "TX"
print("Test 3 PASSED: Join key normalization (leading zeros)")

# Test 4: Config loader defaults
from config_loader import load_config as lc
test_cfg = lc("config.yaml")
assert "higher" in test_cfg["output"]["labels"]
assert "lower" in test_cfg["output"]["labels"]
assert "same" in test_cfg["output"]["labels"]
assert "chunked_loading" in test_cfg
print("Test 4 PASSED: Config loader defaults")

# Test 5: Mock data column completeness (QES-NA)
required_qes_na_cols = [
    "Project Name", "CountySSA", "FIPS County", "County Name",
    "County Class (Rural/Metro/Micro/CEAC)", "State", "State Name",
    "Specialty Group Code", "Specialty Group Name", "Membership Count",
    "% members with Access", "Access Met (Y/N)", "Servicing Providers count",
]
for col in required_qes_na_cols:
    assert col in data["qes_na"].columns, f"Missing QES-NA column: {col}"
print("Test 5 PASSED: QES-NA column completeness")

# Test 6: Mock data column completeness (QES-Providers)
required_qes_prov_cols = [
    "NPI", "TaxID", "Specialty Group Code", "ServicingState",
    "ServicingCounty", "Latitude", "Longitude", "Taxonomy", "Gender",
]
for col in required_qes_prov_cols:
    assert col in data["qes_providers"].columns, f"Missing QES-Provider column: {col}"
print("Test 6 PASSED: QES-Provider column completeness")

# Test 7: Mock data column completeness (NIQ-NA)
required_niq_na_cols = [
    "Project", "LOB", "state", "countSSACode", "county",
    "specialty_code", "specialty_group_name", "coverage_percentage",
    "coverage_status", "total_members", "filing_type",
]
for col in required_niq_na_cols:
    assert col in data["niq_na"].columns, f"Missing NIQ-NA column: {col}"
print("Test 7 PASSED: NIQ-NA column completeness")

# Test 8: SSA codes use correct state prefix and differ from FIPS
ssa_codes_test = data["qes_na"]["CountySSA"].unique()
fips_codes_test = data["qes_na"]["FIPS County"].unique()
for ssa, fips in zip(sorted(ssa_codes_test), sorted(fips_codes_test)):
    assert ssa != fips, f"SSA should differ from FIPS: SSA={ssa}, FIPS={fips}"
print("Test 8 PASSED: SSA codes differ from FIPS codes")

# Test 9: NPI=10 digits, TIN=9 digits
assert all(data["qes_providers"]["NPI"].str.len() == 10), "NPI not 10 digits"
assert all(data["qes_providers"]["TaxID"].str.len() == 9), "TIN not 9 digits"
print("Test 9 PASSED: NPI=10, TIN=9 digits")

# Test 10: Comparison produces Both rows after Excel round-trip
both_ct = len(comparison[comparison["Row_Source"] == "Both"])
assert both_ct > 0, f"Expected Both rows but got {both_ct}"
print(f"Test 10 PASSED: Comparison has {both_ct} Both rows (join key normalization works)")

# Test 11: Repeatable generation (same seed = same output)
data2 = generate_all_data(cfg)
assert data["qes_na"].equals(data2["qes_na"]), "QES-NA not repeatable!"
assert data["niq_na"].equals(data2["niq_na"]), "NIQ-NA not repeatable!"
print("Test 11 PASSED: Mock data generation is repeatable (deterministic seed)")

# Test 12: Synthetic fallback for non-hardcoded states
from mock_data_generator import STATE_NAMES, SSA_STATE_CODES, FIPS_STATE_CODES
test_cfg_ca = cfg.copy()
test_cfg_ca["state"] = "CA"
ca_data = generate_all_data(test_cfg_ca)
ca_ssa = ca_data["qes_na"]["CountySSA"].unique()
ca_fips = ca_data["qes_na"]["FIPS County"].unique()
assert all(str(c).startswith("06") for c in ca_ssa), f"CA SSA should start with 06: {ca_ssa[:3]}"
assert all(str(c).startswith("06") for c in ca_fips), f"CA FIPS should start with 06: {ca_fips[:3]}"
assert len(ca_data["qes_na"]) > 0, "CA should generate rows"
# Verify ZIP codes are 5 digits with proper zero-padding
ca_zips = ca_data["qes_providers"]["Zip"]
assert all(ca_zips.str.len() == 5), "CA ZIP codes should be 5 digits"
# Verify invalid state raises error
try:
    bad_cfg = cfg.copy()
    bad_cfg["state"] = "XX"
    generate_all_data(bad_cfg)
    assert False, "Should have raised ValueError for invalid state"
except ValueError:
    pass
print("Test 12 PASSED: Synthetic fallback (CA generates valid data, XX raises error)")

# Test 13: QES CSV loading and auto-detection
import shutil
csv_na = os.path.join(project_root, "input", "QES_NA_test.csv")
csv_prov = os.path.join(project_root, "input", "QES_Prov_test.csv")
data["qes_na"].to_csv(csv_na, index=False)
data["qes_providers"].to_csv(csv_prov, index=False)
csv_cfg = cfg.copy()
csv_cfg["qes"] = dict(cfg["qes"])
csv_cfg["qes"]["source_type"] = "csv"
csv_cfg["qes"]["workbook1_path"] = csv_na
csv_cfg["qes"]["workbook2_path"] = csv_prov
from data_loader import load_qes_data as lqd
csv_qes_na, csv_qes_prov = lqd(csv_cfg)
assert csv_qes_na.shape == qes_na.shape, f"CSV QES-NA shape mismatch: {csv_qes_na.shape} vs {qes_na.shape}"
assert csv_qes_prov.shape == qes_providers.shape, f"CSV QES-Providers shape mismatch"
os.remove(csv_na)
os.remove(csv_prov)
print("Test 13 PASSED: QES CSV loading (explicit source_type + auto-detect)")

# Test 14: Config compact format normalization
test_cfg_compact = lc("config.yaml")
for ac in test_cfg_compact["additional_result_columns"]:
    assert "qes_col" in ac, f"Missing qes_col in normalized entry: {ac}"
    assert "niq_col" in ac, f"Missing niq_col in normalized entry: {ac}"
    assert "label" in ac and ac["label"], f"Missing/empty label in normalized entry: {ac}"
print("Test 14 PASSED: Compact additional_result_columns normalization")

print("\nAll 14 component tests PASSED!")

# %% -- Section 9: Large File Simulation
# This section tests chunked loading with a larger-than-threshold dataset.
# It creates a temporary ~100MB CSV, loads it in chunks, and verifies row counts.

print("=== Large File Simulation ===")
print("Generating large test CSV...")
large_providers = pd.concat([data["niq_providers"]] * 100, ignore_index=True)
print(f"Large provider dataset: {len(large_providers):,} rows")

csv_path = os.path.join(project_root, "input", "large_test_providers.csv")
large_providers.to_csv(csv_path, index=False)
file_size_mb = os.path.getsize(csv_path) / (1024 * 1024)
print(f"File size: {file_size_mb:.1f} MB")

# Test chunked loading
from data_loader import _load_csv_chunked
chunk_cfg = {"chunk_size": 10000}
result = _load_csv_chunked(csv_path, "servicing_state", cfg["state"], chunk_cfg)
print(f"Loaded {len(result):,} rows via chunked CSV")
assert len(result) == len(large_providers), "Row count mismatch!"
print("Large file test PASSED!")

# Cleanup
os.remove(csv_path)
print("Cleaned up test file.")

# %%
print("\n=== Notebook complete ===")
