"""
Workbook Builder: Assembles the final output Excel workbook.
  - Enhanced Summary dashboard (5 sections, left-aligned)
  - Comparison_Results (with clickable hyperlinks on Difference column)
  - Drill-down sub-sheets per County x Specialty (QES and NIQ cross-linked)
  - QES_Network_Adequacy, QES_Providers, NIQ_Network_Adequacy, NIQ_Providers
All labels are plain ASCII from config.
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def build_output_workbook(
    qes_na: pd.DataFrame,
    qes_providers: pd.DataFrame,
    niq_na: pd.DataFrame,
    niq_providers: pd.DataFrame,
    comparison: pd.DataFrame,
    cfg: dict,
) -> str:
    out_cfg = cfg["output"]
    colors = out_cfg.get("colors", {})
    labels = out_cfg["labels"]

    wb = Workbook()
    wb.remove(wb.active)

    styles = _build_styles(colors)

    # --- 1. Create drill-down sub-sheets ---
    key_qes = cfg["key_columns"]["qes"]
    drilldown_cfg = cfg.get("drilldown", {})
    max_dd = out_cfg.get("max_drilldown_sheets", 200)

    # Determine provider filter keys from drilldown config
    qes_prov_keys = drilldown_cfg.get("qes_provider_keys", {})
    niq_prov_keys = drilldown_cfg.get("niq_provider_keys", {})

    # Use key columns to identify county and specialty in comparison results
    county_col = key_qes[1]   # CountySSA
    spec_col = key_qes[2]     # Specialty Group Code

    # For provider filtering, we need the county name (not SSA) for display
    # but use the configured provider keys for filtering
    qes_prov_county_col = qes_prov_keys.get("county", "ServicingCounty")
    qes_prov_spec_col = qes_prov_keys.get("specialty", "Specialty Group Code")
    niq_prov_county_col = niq_prov_keys.get("county", "servicing_county")
    niq_prov_spec_col = niq_prov_keys.get("specialty", "specialty_group_code")

    # Build list of unique county+specialty combos from comparison
    combos = []
    for _, row in comparison.iterrows():
        combo = (row.get(county_col, ""), row.get(spec_col, ""))
        if combo not in combos:
            combos.append(combo)

    # We need county name for display in sheet names. Try to get from additional cols.
    # Build a lookup from SSA -> County Name from comparison data
    county_name_col = None
    for ac in cfg.get("additional_result_columns", []):
        if ac["label"] == "County Name":
            county_name_col = f"QES_{ac['label']}"
            break

    ssa_to_name = {}
    if county_name_col and county_name_col in comparison.columns:
        for _, row in comparison.iterrows():
            ssa = row.get(county_col, "")
            name = row.get(county_name_col, "")
            if ssa and name and pd.notna(name):
                ssa_to_name[str(ssa)] = str(name)

    qes_dd_index = {}
    niq_dd_index = {}
    dd_count = 0
    all_dd_names = []

    for county_ssa, specialty in combos:
        if dd_count >= max_dd:
            break

        county_display = ssa_to_name.get(str(county_ssa), str(county_ssa))

        # QES drill-down -- filter providers by county and specialty
        qes_filtered = _filter_providers(
            qes_providers, qes_prov_county_col, county_display, qes_prov_spec_col, specialty
        )
        # Also try filtering by SSA if county name didn't match
        if len(qes_filtered) == 0 and county_display != str(county_ssa):
            qes_filtered = _filter_providers(
                qes_providers, qes_prov_county_col, str(county_ssa), qes_prov_spec_col, specialty
            )

        if len(qes_filtered) > 0:
            sname = _safe_sheet_name(f"QES_{county_display}_{specialty}", all_dd_names)
            all_dd_names.append(sname)
            ws = wb.create_sheet(title=sname)
            qes_dd_index[(county_ssa, specialty)] = sname
            dd_count += 1
        else:
            sname = None

        # NIQ drill-down
        niq_filtered = _filter_providers(
            niq_providers, niq_prov_county_col, county_display, niq_prov_spec_col, specialty
        )
        if len(niq_filtered) == 0 and county_display != str(county_ssa):
            niq_filtered = _filter_providers(
                niq_providers, niq_prov_county_col, str(county_ssa), niq_prov_spec_col, specialty
            )

        niq_sname = None
        if len(niq_filtered) > 0:
            niq_sname = _safe_sheet_name(f"NIQ_{county_display}_{specialty}", all_dd_names)
            all_dd_names.append(niq_sname)
            ws_niq = wb.create_sheet(title=niq_sname)
            niq_dd_index[(county_ssa, specialty)] = niq_sname
            dd_count += 1

        # Write the sheets with cross-links
        if (county_ssa, specialty) in qes_dd_index:
            _write_drilldown_sheet(
                wb[qes_dd_index[(county_ssa, specialty)]],
                qes_filtered, county_display, specialty, "QES",
                niq_sname, styles, out_cfg
            )
        if niq_sname:
            _write_drilldown_sheet(
                wb[niq_sname],
                niq_filtered, county_display, specialty, "NIQ",
                qes_dd_index.get((county_ssa, specialty)), styles, out_cfg
            )

    print(f"  [built] {dd_count} drill-down sheets")

    # --- 2. Comparison Results ---
    ws_comp = wb.create_sheet(title="Comparison_Results", index=0)
    _write_comparison_sheet(ws_comp, comparison, cfg, qes_dd_index, niq_dd_index, styles, labels, out_cfg)

    # --- 3. Summary ---
    ws_summ = wb.create_sheet(title="Summary", index=0)
    _write_summary_sheet(ws_summ, comparison, qes_na, niq_na, qes_providers, niq_providers, cfg, styles, labels)

    # --- 4. Raw data sheets ---
    large_threshold = 50000
    for title, df in [
        ("QES_Network_Adequacy", qes_na),
        ("QES_Providers", qes_providers),
        ("NIQ_Network_Adequacy", niq_na),
        ("NIQ_Providers", niq_providers),
    ]:
        ws = wb.create_sheet(title=title)
        if len(df) > large_threshold:
            _write_data_sheet_chunked(ws, df, styles, out_cfg)
        else:
            _write_data_sheet(ws, df, styles, out_cfg)

    # --- Save (with state and timestamp in filename) ---
    base_path = out_cfg["workbook_path"]
    root, ext = os.path.splitext(base_path)
    state = cfg.get("state", "XX")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_path = f"{root}_{state}_{timestamp}{ext}"
    os.makedirs(os.path.dirname(final_path) or ".", exist_ok=True)
    wb.save(final_path)
    print(f"\n  [saved] {final_path} ({len(wb.sheetnames)} sheets)")
    return final_path


# ============================================================================
# Styles
# ============================================================================

def _build_styles(colors: dict) -> dict:
    return {
        "header_fill": PatternFill("solid", fgColor=colors.get("header_fill", "1F4E79")),
        "header_font": Font(name="Arial", bold=True, color=colors.get("header_font", "FFFFFF"), size=10),
        "data_font": Font(name="Arial", size=10),
        "match_fill": PatternFill("solid", fgColor=colors.get("match_fill", "C6EFCE")),
        "mismatch_fill": PatternFill("solid", fgColor=colors.get("mismatch_fill", "FFC7CE")),
        "qes_only_fill": PatternFill("solid", fgColor=colors.get("qes_only_fill", "FCE4D6")),
        "niq_only_fill": PatternFill("solid", fgColor=colors.get("niq_only_fill", "D6E4FC")),
        "hyperlink_font": Font(name="Arial", size=10, underline="single",
                               color=colors.get("hyperlink_font", "0563C1")),
        "border": Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        ),
        "title_font": Font(name="Arial", bold=True, size=14, color="1F4E79"),
        "subtitle_font": Font(name="Arial", bold=True, size=11, color="1F4E79"),
        "metric_font": Font(name="Arial", size=11),
        "big_number_font": Font(name="Arial", bold=True, size=16, color="1F4E79"),
        "bold_font": Font(name="Arial", bold=True, size=11),
        "section_header_fill": PatternFill("solid", fgColor="D6DCE4"),
    }


# ============================================================================
# Sheet Writers
# ============================================================================

def _write_data_sheet(ws, df, styles, out_cfg):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = styles["border"]
            if r_idx == 1:
                cell.fill = styles["header_fill"]
                cell.font = styles["header_font"]
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            else:
                cell.font = styles["data_font"]
    if out_cfg.get("freeze_panes"):
        ws.freeze_panes = "A2"
    if out_cfg.get("auto_column_width"):
        _auto_width(ws)


def _write_data_sheet_chunked(ws, df, styles, out_cfg, chunk_size=10000):
    """Write a large DataFrame to a worksheet in chunks."""
    # Write header
    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = styles["border"]

    # Write data in chunks
    total_rows = len(df)
    for start in range(0, total_rows, chunk_size):
        end = min(start + chunk_size, total_rows)
        chunk = df.iloc[start:end]
        for r_offset, (_, data_row) in enumerate(chunk.iterrows()):
            r_idx = start + r_offset + 2
            for c_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=data_row[col_name])
                cell.font = styles["data_font"]
                cell.border = styles["border"]
        if total_rows > chunk_size:
            print(f"    ... wrote rows {start+1:,} to {end:,} of {total_rows:,}")

    if out_cfg.get("freeze_panes"):
        ws.freeze_panes = "A2"
    # Skip auto_column_width for large sheets (performance)
    if total_rows < 50000 and out_cfg.get("auto_column_width"):
        _auto_width(ws)


def _write_comparison_sheet(ws, comparison, cfg, qes_dd_index, niq_dd_index, styles, labels, out_cfg):
    drilldown_cfg = cfg.get("drilldown", {})
    key_qes = cfg["key_columns"]["qes"]
    county_col = key_qes[1]
    spec_col = key_qes[2]

    # Identify the drill-down link column (Diff column)
    dd_link_col = drilldown_cfg.get("link_column", None)

    columns = list(comparison.columns)

    # Header
    for c_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = styles["border"]

    # Data rows
    for r_idx, (_, data_row) in enumerate(comparison.iterrows(), 2):
        county = data_row.get(county_col, "")
        specialty = data_row.get(spec_col, "")
        row_source = data_row.get("Row_Source", "Both")

        for c_idx, col_name in enumerate(columns, 1):
            value = data_row[col_name]
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = styles["border"]
            cell.font = styles["data_font"]

            # Hyperlink on Diff column -> QES drill-down sheet
            if col_name == dd_link_col and (county, specialty) in qes_dd_index:
                cell.value = value if not pd.isna(value) else ""
                cell.hyperlink = f"#'{qes_dd_index[(county, specialty)]}'!A1"
                cell.font = styles["hyperlink_font"]
                continue

            cell.value = "" if pd.isna(value) else value

            # Conditional coloring
            if col_name.startswith("Match_") or col_name == "Overall_Match":
                str_val = str(value)
                if str_val == labels["match"] or str_val == labels["overall_match"]:
                    cell.fill = styles["match_fill"]
                elif str_val == labels["mismatch"] or str_val == labels["overall_mismatch"]:
                    cell.fill = styles["mismatch_fill"]
                elif str_val == labels["warning"]:
                    if row_source == "QES Only":
                        cell.fill = styles["qes_only_fill"]
                    else:
                        cell.fill = styles["niq_only_fill"]
                elif str_val in (labels["overall_qes_only"], labels["overall_niq_only"]):
                    cell.fill = styles["qes_only_fill"] if "QES" in str_val else styles["niq_only_fill"]

            # Color Diff columns -- RED when deviating from QES (master)
            if col_name.startswith("Diff_"):
                str_val = str(value) if not pd.isna(value) else ""
                if str_val and str_val != "0" and str_val != "0.0" and str_val != "":
                    if str_val not in (labels.get("na_qes_only", ""), labels.get("na_niq_only", "")):
                        cell.fill = styles["mismatch_fill"]

            # Color direction cells -- QES is master, deviations are RED
            if col_name.startswith("Direction_"):
                str_val = str(value)
                if str_val in (labels.get("higher", "HIGHER"), labels.get("lower", "LOWER")):
                    cell.fill = styles["mismatch_fill"]  # RED - deviation from QES
                elif str_val == labels.get("same", "SAME"):
                    cell.fill = styles["match_fill"]

    if out_cfg.get("freeze_panes"):
        ws.freeze_panes = "A2"
    if out_cfg.get("auto_column_width"):
        _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


def _write_drilldown_sheet(ws, df, county, specialty, source_label, other_sheet_name, styles, out_cfg):
    title = f"{source_label} Providers -- {county} / {specialty}"
    ws.cell(row=1, column=1, value=title).font = styles["title_font"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(df.columns), 9))

    # Back link
    back_cell = ws.cell(row=2, column=1, value="<< Back to Comparison Results")
    back_cell.hyperlink = "#Comparison_Results!A1"
    back_cell.font = styles["hyperlink_font"]

    # Cross-link to other dataset's provider sheet
    if other_sheet_name:
        other_label = "NIQ" if source_label == "QES" else "QES"
        other_cell = ws.cell(row=2, column=4, value=f">> See {other_label} Providers")
        other_cell.hyperlink = f"#'{other_sheet_name}'!A1"
        other_cell.font = styles["hyperlink_font"]

    # Provider count
    ws.cell(row=3, column=1, value=f"Total Providers: {len(df)}").font = styles["bold_font"]

    start_row = 5
    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = styles["border"]

    for r_idx, (_, data_row) in enumerate(df.iterrows(), start_row + 1):
        for c_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=data_row[col_name])
            cell.font = styles["data_font"]
            cell.border = styles["border"]

    if out_cfg.get("freeze_panes"):
        ws.freeze_panes = f"A{start_row + 1}"
    if out_cfg.get("auto_column_width"):
        _auto_width(ws)


# ============================================================================
# Summary Sheet (5 Sections)
# ============================================================================

def _write_summary_sheet(ws, comparison, qes_na, niq_na, qes_providers, niq_providers, cfg, styles, labels):
    """Build enhanced summary with 5 sections, all left-aligned."""
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20

    r = 2
    ws.cell(row=r, column=2, value="Network Adequacy Comparison Report").font = styles["title_font"]
    r += 1
    ws.cell(row=r, column=2, value=f"State: {cfg['state']}").font = styles["subtitle_font"]
    r += 1
    ws.cell(row=r, column=2, value=f"NIQ Source: {cfg['niq']['source_type']}").font = styles["metric_font"]
    r += 2

    # Overall metrics
    total = len(comparison)
    both = len(comparison[comparison["Row_Source"] == "Both"])
    matches = len(comparison[comparison["Overall_Match"] == labels["overall_match"]])
    mismatches = len(comparison[comparison["Overall_Match"] == labels["overall_mismatch"]])
    qes_only = len(comparison[comparison["Row_Source"] == "QES Only"])
    niq_only = len(comparison[comparison["Row_Source"] == "NIQ Only"])
    match_pct = f"{matches/both*100:.1f}%" if both > 0 else "N/A"

    ws.cell(row=r, column=2, value="Overall Metrics").font = styles["subtitle_font"]
    r += 1
    metrics = [
        ("Total County x Specialty Combinations", total),
        ("Present in Both Systems", both),
        ("Fully Matching", matches),
        ("With Differences", mismatches),
        ("In QES Only (not in NIQ)", qes_only),
        ("In NIQ Only (not in QES)", niq_only),
        ("Match Rate", match_pct),
    ]
    for label_text, val in metrics:
        ws.cell(row=r, column=2, value=label_text).font = styles["metric_font"]
        c = ws.cell(row=r, column=3, value=val)
        c.font = styles["big_number_font"] if label_text == "Match Rate" else styles["bold_font"]
        c.alignment = Alignment(horizontal="left")
        r += 1

    r += 1

    # --- Section i: Number of Counties ---
    r = _write_county_count_section(ws, r, comparison, cfg, styles)

    # --- Section ii: Unique Provider Count by Distinct NPI per State ---
    r = _write_unique_npi_section(ws, r, qes_providers, niq_providers, cfg, styles)

    # --- Section iii: Summary by Specialty ---
    r = _write_specialty_summary_section(ws, r, comparison, cfg, styles, labels)

    # --- Section iv: Zero Serving Providers ---
    r = _write_zero_serving_section(ws, r, qes_na, niq_na, cfg, styles)

    # --- Section v: Mismatches by Column + Link ---
    r += 1
    ws.cell(row=r, column=2, value="v) Mismatches by Column").font = styles["subtitle_font"]
    r += 1
    for cc in cfg["compare_columns"]:
        match_col = f"Match_{cc['label']}"
        if match_col in comparison.columns:
            mm_count = len(comparison[comparison[match_col] == labels["mismatch"]])
            ws.cell(row=r, column=2, value=cc["label"]).font = styles["metric_font"]
            ws.cell(row=r, column=3, value=mm_count).font = styles["bold_font"]
            r += 1

    r += 1
    link_cell = ws.cell(row=r, column=2, value=">> View Full Comparison Details")
    link_cell.hyperlink = "#Comparison_Results!A1"
    link_cell.font = styles["hyperlink_font"]


def _write_county_count_section(ws, r, comparison, cfg, styles):
    key_qes = cfg["key_columns"]["qes"]
    county_col = key_qes[1]

    ws.cell(row=r, column=2, value="i) Number of Counties").font = styles["subtitle_font"]
    r += 1
    unique_counties = comparison[county_col].nunique() if county_col in comparison.columns else 0
    ws.cell(row=r, column=2, value="Total Unique Counties (by SSA Code)").font = styles["metric_font"]
    ws.cell(row=r, column=3, value=unique_counties).font = styles["big_number_font"]
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="left")
    r += 2
    return r


def _write_unique_npi_section(ws, r, qes_providers, niq_providers, cfg, styles):
    summary_cfg = cfg.get("summary", {})
    npi_cols = summary_cfg.get("npi_columns", {"qes": "NPI", "niq": "npi"})

    ws.cell(row=r, column=2, value="ii) Unique Provider Count (Distinct NPI per State)").font = styles["subtitle_font"]
    r += 1

    # Headers
    for col_idx, header in enumerate(["Source", "State", "Distinct NPI Count"], 2):
        cell = ws.cell(row=r, column=col_idx, value=header)
        cell.font = styles["bold_font"]
        cell.fill = styles["section_header_fill"]
        cell.border = styles["border"]
    r += 1

    # QES NPI count
    qes_npi_col = npi_cols.get("qes", "NPI")
    if qes_npi_col in qes_providers.columns:
        qes_npi_count = qes_providers[qes_npi_col].nunique()
        ws.cell(row=r, column=2, value="QES").font = styles["metric_font"]
        ws.cell(row=r, column=3, value=cfg["state"]).font = styles["metric_font"]
        ws.cell(row=r, column=4, value=qes_npi_count).font = styles["bold_font"]
        r += 1

    # NIQ NPI count
    niq_npi_col = npi_cols.get("niq", "npi")
    if niq_npi_col in niq_providers.columns:
        niq_npi_count = niq_providers[niq_npi_col].nunique()
        ws.cell(row=r, column=2, value="NIQ").font = styles["metric_font"]
        ws.cell(row=r, column=3, value=cfg["state"]).font = styles["metric_font"]
        ws.cell(row=r, column=4, value=niq_npi_count).font = styles["bold_font"]
        r += 1

    r += 1
    return r


def _write_specialty_summary_section(ws, r, comparison, cfg, styles, labels):
    key_qes = cfg["key_columns"]["qes"]
    spec_col = key_qes[2]
    summary_cfg = cfg.get("summary", {})
    spec_desc_cols = summary_cfg.get("specialty_desc_columns", {})
    qes_desc_col = spec_desc_cols.get("qes")

    # Find direction column
    dir_col = None
    for cc in cfg["compare_columns"]:
        if cc.get("direction_indicator"):
            dir_col = f"Direction_{cc['label']}"
            break

    ws.cell(row=r, column=2, value="iii) Summary by Specialty").font = styles["subtitle_font"]
    r += 1

    # Table headers
    headers = ["Specialty Code", "Specialty Description", "Matching Counties", "QES > NIQ", "NIQ > QES"]
    for col_idx, header in enumerate(headers, 2):
        cell = ws.cell(row=r, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="left")
        cell.border = styles["border"]
    r += 1

    both_rows = comparison[comparison["Row_Source"] == "Both"]

    # Get spec description from additional columns if available
    qes_spec_desc_result = None
    if qes_desc_col:
        qes_spec_desc_result = f"QES_{qes_desc_col}" if f"QES_{qes_desc_col}" not in comparison.columns else None
        # Try the label-based column name
        for ac in cfg.get("additional_result_columns", []):
            if ac.get("qes_col") == qes_desc_col or ac["label"] == "Specialty Group Name":
                qes_spec_desc_result = f"QES_{ac['label']}"
                break

    if spec_col not in both_rows.columns:
        r += 1
        return r

    for spec_code in sorted(both_rows[spec_col].unique()):
        spec_rows = both_rows[both_rows[spec_col] == spec_code]

        # Get description
        spec_desc = ""
        if qes_spec_desc_result and qes_spec_desc_result in spec_rows.columns:
            descs = spec_rows[qes_spec_desc_result].dropna().unique()
            if len(descs) > 0:
                spec_desc = str(descs[0])

        if dir_col and dir_col in spec_rows.columns:
            matching = len(spec_rows[spec_rows[dir_col] == labels.get("same", "SAME")])
            # QES > NIQ means NIQ is LOWER
            qes_higher = len(spec_rows[spec_rows[dir_col] == labels.get("lower", "LOWER")])
            # NIQ > QES means NIQ is HIGHER
            niq_higher = len(spec_rows[spec_rows[dir_col] == labels.get("higher", "HIGHER")])
        else:
            matching = len(spec_rows[spec_rows["Overall_Match"] == labels["overall_match"]])
            qes_higher = 0
            niq_higher = 0

        ws.cell(row=r, column=2, value=spec_code).font = styles["metric_font"]
        ws.cell(row=r, column=2).border = styles["border"]
        ws.cell(row=r, column=3, value=spec_desc).font = styles["metric_font"]
        ws.cell(row=r, column=3).border = styles["border"]
        cell_match = ws.cell(row=r, column=4, value=matching)
        cell_match.font = styles["bold_font"]
        cell_match.fill = styles["match_fill"]
        cell_match.border = styles["border"]
        cell_qes = ws.cell(row=r, column=5, value=qes_higher)
        cell_qes.font = styles["bold_font"]
        cell_qes.fill = styles["qes_only_fill"]
        cell_qes.border = styles["border"]
        cell_niq = ws.cell(row=r, column=6, value=niq_higher)
        cell_niq.font = styles["bold_font"]
        cell_niq.fill = styles["niq_only_fill"]
        cell_niq.border = styles["border"]
        r += 1

    r += 1
    return r


def _write_zero_serving_section(ws, r, qes_na, niq_na, cfg, styles):
    summary_cfg = cfg.get("summary", {})
    serving_cols = summary_cfg.get("serving_columns", {})
    qes_serving = serving_cols.get("qes", "Servicing Providers count")
    niq_serving = serving_cols.get("niq", "provider_covering")

    ws.cell(row=r, column=2, value="iv) Counties with 0 Serving Providers").font = styles["subtitle_font"]
    r += 1

    qes_zero = 0
    if qes_serving in qes_na.columns:
        qes_zero = len(qes_na[qes_na[qes_serving] == 0])

    niq_zero = 0
    if niq_serving in niq_na.columns:
        niq_zero = len(niq_na[niq_na[niq_serving] == 0])

    ws.cell(row=r, column=2, value="QES - County x Specialty with 0 Serving Providers").font = styles["metric_font"]
    ws.cell(row=r, column=3, value=qes_zero).font = styles["bold_font"]
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="left")
    r += 1
    ws.cell(row=r, column=2, value="NIQ - County x Specialty with 0 Serving Providers").font = styles["metric_font"]
    ws.cell(row=r, column=3, value=niq_zero).font = styles["bold_font"]
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="left")
    r += 2
    return r


# ============================================================================
# Utilities
# ============================================================================

def _safe_sheet_name(name: str, existing: list, max_len: int = 31) -> str:
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, "_")
    name = name[:max_len]
    original = name
    counter = 1
    while name in existing:
        suffix = f"_{counter}"
        name = original[:max_len - len(suffix)] + suffix
        counter += 1
    return name


def _filter_providers(df, county_col, county, spec_col, specialty):
    if county_col not in df.columns or spec_col not in df.columns:
        return df.head(0)
    mask = (df[county_col].astype(str) == str(county)) & (df[spec_col].astype(str) == str(specialty))
    return df[mask].reset_index(drop=True)


def _auto_width(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)
